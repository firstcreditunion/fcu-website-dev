"""Generates supabase/migrations/0002_project_hub_seed.sql from the FCU timeline workbook.
Run once: python scripts/db/generate-seed.py "C:\\Users\\ivicliph\\Downloads\\FCU_Website_Project_Timeline_2026.xlsx"
Normalization rules: docs/superpowers/specs/2026-06-12-project-hub-design.md
"""
import sys, uuid, datetime, re
from openpyxl import load_workbook

XLSX = sys.argv[1]
OUT = "supabase/migrations/0002_project_hub_seed.sql"
SLUG = "fcu-website-rebuild-2026"
YEAR = 2026
MONTHS = {m: i + 1 for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"])}
PHASE_COLORS = {1: "chart-1", 2: "chart-2", 3: "chart-3", 4: "chart-4", 5: "chart-5", 6: "fcu-mint-500"}
STATUS_MAP = {"complete": "complete", "underway": "in_progress", "n/a": "na", "": "not_started"}

def esc(s):
    return str(s).replace("'", "''").strip()

def parse_date(v):
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    m = re.match(r"^(\d{1,2})\s+([A-Za-z]{3})", str(v).strip())
    if m:
        return datetime.date(YEAR, MONTHS[m.group(2).title()], int(m.group(1))).isoformat()
    raise ValueError(f"unparseable date: {v!r}")

def sqld(d):
    return f"'{d}'" if d else "null"

def fix_typos(s):
    return s.replace("rotateble", "rotatable").replace("rotabele", "rotatable")

def networkdays(a, b):
    a, b = datetime.date.fromisoformat(a), datetime.date.fromisoformat(b)
    d, n = a, 0
    while d <= b:
        if d.weekday() < 5:
            n += 1
        d += datetime.timedelta(days=1)
    return n

wb = load_workbook(XLSX, data_only=True)
ws = wb["Project Summary"]
rows = [[c.value for c in r] for r in ws.iter_rows()]

def cell(r, c):
    v = rows[r - 1][c - 1] if r - 1 < len(rows) and c - 1 < len(rows[r - 1]) else None
    return "" if v is None else str(v).strip()

pid = str(uuid.uuid4())
stmts = []
stmts.append(
    f"insert into api.pt_projects (id, slug, title, subtitle) values "
    f"('{pid}', '{SLUG}', '{esc(cell(4, 2))}', "
    f"'7 months (Feb – Aug 2026) · Next.js 16 · Supabase · Sanity · Vercel');"
)

def ins(table, cols, vals_rows):
    for v in vals_rows:
        stmts.append(f"insert into api.{table} ({cols}) values ({v});")

# facts: overview rows 4-7, deployment rows 55-62; strip trailing ':' from labels
facts = []
for i, r in enumerate(range(4, 8)):
    facts.append(f"'{pid}', 'overview', '{esc(cell(r, 1).rstrip(':'))}', '{esc(cell(r, 2))}', {i}")
for i, r in enumerate(range(55, 63)):
    facts.append(f"'{pid}', 'deployment', '{esc(cell(r, 1).rstrip(':'))}', '{esc(cell(r, 2))}', {i}")
ins("pt_facts", "project_id, section, label, value, sort_order", facts)

# tech stack rows 11-20
ins("pt_tech_stack", "project_id, technology, version, purpose, sort_order",
    [f"'{pid}', '{esc(cell(r, 1))}', '{esc(cell(r, 2))}', '{esc(cell(r, 3))}', {i}"
     for i, r in enumerate(range(11, 21))])

# phases rows 25-30 (number, name, duration, start, end)
phase_ids = {}
phase_rows = []
for r in range(25, 31):
    num = int(cell(r, 1))
    phase_ids[num] = str(uuid.uuid4())
    phase_rows.append(
        f"'{phase_ids[num]}', '{pid}', {num}, '{esc(cell(r, 2))}', "
        f"{sqld(parse_date(rows[r - 1][3]))}, {sqld(parse_date(rows[r - 1][4]))}, '{PHASE_COLORS[num]}'")
ins("pt_phases", "id, project_id, phase_number, name, start_date, end_date, color_token", phase_rows)

# milestones rows 35-41
ins("pt_milestones", "project_id, date, name, deliverable, sort_order",
    [f"'{pid}', {sqld(parse_date(rows[r - 1][0]))}, '{esc(cell(r, 2))}', '{esc(cell(r, 3))}', {i}"
     for i, r in enumerate(range(35, 42))])

# risks rows 46-51
ins("pt_risks", "project_id, risk, impact, mitigation, sort_order",
    [f"'{pid}', '{esc(cell(r, 1))}', '{cell(r, 2).lower()}', '{esc(cell(r, 3))}', {i}"
     for i, r in enumerate(range(46, 52))])

# deliverables rows 66-80 (strip bullet, fix typos)
ins("pt_deliverables", "project_id, description, sort_order",
    [f"'{pid}', '{esc(fix_typos(cell(r, 1).lstrip('•').strip()))}', {i}"
     for i, r in enumerate(range(66, 81)) if cell(r, 1)])

# timeline sheet: phase headers (CAPS), groups (no dates), tasks (indented, dates)
ws2 = wb["Project Timeline"]
trows = [[c.value for c in r] for r in ws2.iter_rows()]
group_id = None
group_sort = {}
task_sort = {}
gstmts, tstmts = [], []
for raw in trows[4:]:  # data starts row 5
    pcol = "" if raw[0] is None else str(raw[0]).strip()
    name = "" if len(raw) < 2 or raw[1] is None else str(raw[1])
    if not pcol.startswith("Phase") or not name.strip():
        continue
    pnum = int(pcol.split()[1])
    duration = "" if len(raw) < 3 or raw[2] is None else str(raw[2]).strip()
    start = raw[3] if len(raw) > 3 else None
    end = raw[4] if len(raw) > 4 else None
    status_raw = ("" if len(raw) < 13 or raw[12] is None else str(raw[12]).strip().lower())
    if name.strip().isupper():
        continue  # phase header row; phases come from the summary sheet
    if not duration and start is None and end is None:
        group_id = str(uuid.uuid4())
        s = group_sort.get(pnum, 0)
        group_sort[pnum] = s + 1
        gstmts.append(
            f"insert into api.pt_task_groups (id, project_id, phase_id, name, sort_order) values "
            f"('{group_id}', '{pid}', '{phase_ids[pnum]}', '{esc(name)}', {s});")
        continue
    if group_id is None:
        raise ValueError(f"task before any group: {name!r}")
    # keep the label only when NOT derivable from the dates: 'Ongoing' stays;
    # '1 week' matching the span is dropped (UI computes it); a label that
    # CONTRADICTS the dates stays visible (spec: seeded as-entered).
    dl = duration if duration else None
    sd, ed = parse_date(start), parse_date(end)
    if dl:
        m = re.fullmatch(r"(\d+) (week|weeks|day|days)", dl)
        if m and sd and ed:
            n, unit = int(m.group(1)), m.group(2)
            wd = networkdays(sd, ed)
            if (unit.startswith("day") and wd == n) or (unit.startswith("week") and wd == n * 5):
                dl = None
    s = task_sort.get(group_id, 0)
    task_sort[group_id] = s + 1
    status = STATUS_MAP.get(status_raw, "not_started")
    dl_sql = f"'{esc(dl)}'" if dl else "null"
    tstmts.append(
        f"insert into api.pt_tasks (project_id, group_id, name, start_date, end_date, duration_label, status, sort_order) values "
        f"('{pid}', '{group_id}', '{esc(name)}', {sqld(sd)}, {sqld(ed)}, {dl_sql}, '{status}', {s});")

body = "\n  ".join(s.replace("\n", " ") for s in stmts + gstmts + tstmts)
sql = f"""-- Generated by scripts/db/generate-seed.py — do not hand-edit row values.
do $$
begin
  if exists (select 1 from api.pt_projects where slug = '{SLUG}') then
    raise notice 'project hub already seeded; skipping';
    return;
  end if;
  {body}
end $$;
"""
with open(OUT, "w", encoding="utf-8") as f:
    f.write(sql)
print(f"wrote {OUT}: phases=6, groups={sum(group_sort.values())}, tasks={len(tstmts)}")
